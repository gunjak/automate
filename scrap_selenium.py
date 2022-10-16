from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl 
from function import append_data_to_excel
from openpyxl import load_workbook
import os


#creating workbook
EXCEL_PATH='C:\\Users\\Gunja\\Desktop\\web_scrapping\\company_data.csv'
if os.path.isfile(EXCEL_PATH):
    wbook=load_workbook(EXCEL_PATH)
    sheet=wbook.get_active_sheet()
else:
    wbook = openpyxl.Workbook() 
    sheet=wbook.active
    

sheet.append(['Company','Website','Location','Contact','Rating','Riview Count','Hourly Rate','Mini Project Size','Employee Size'])

#creating main webdriver
URL='https://clutch.co/'
DRIVER_PATH = r'C:\Users\Gunja\Downloads\chromedriver_win32 (3)\chromedriver'
wb=webdriver.Chrome(executable_path=DRIVER_PATH)
wb.get(URL)



for domain_class in wb.find_elements(By.CLASS_NAME,'sitemap-nav__item'):
    count=0
    domain_url=domain_class.get_attribute('href')
    #creating driver for specific domain 
    while True:
        if count==0:
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.get(domain_url)
            last_page_url=driver.find_element(By.XPATH,'//*[@id="providers"]/nav/ul/li[13]/a').get_attribute('href')
        else:
            domain_url=domain_url+'?page='+str(count)
            #creating driver to next page domain
            driver = webdriver.Chrome(ChromeDriverManager().install())
            driver.get(domain_url)
        for company_class in driver.find_elements(By.CLASS_NAME,'company_info'):
            company_url=company_class.find_element(By.TAG_NAME,'a').get_attribute('href')
            print(company_url)
            print(domain_url)
            data=append_data_to_excel(company_url)
            sheet.append(data)
            wbook.save(EXCEL_PATH)
        driver.quit()
        if domain_url==last_page_url:
            break
        count+=1   
wb.quit()    